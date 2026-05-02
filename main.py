from fastapi import FastAPI, Request, UploadFile
from starlette.datastructures import UploadFile as StarletteUploadFile
from openpyxl import load_workbook
from io import BytesIO
from datetime import datetime
import re

app = FastAPI()


def parse_amount(value):
    if value is None:
        return 0

    text = str(value).strip()
    text = text.replace("TL", "").replace("₺", "").strip()
    text = text.replace(".", "").replace(",", ".")

    try:
        return float(text)
    except Exception:
        return 0


def clean_date(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")

    text = str(value).strip()

    for fmt in ["%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d"]:
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except Exception:
            pass

    return text


def extract_field(text, labels):
    if not text:
        return ""

    for label in labels:
        pattern = label + r"\s*:?\s*(.*)"
        match = re.search(pattern, text, re.IGNORECASE)

        if match:
            value = match.group(1).strip()

            stop_words = [
                "IBAN:",
                "Açıklama:",
                "Tutar:",
                "Referans:",
                "Gönderen:",
                "Alıcı:",
                "Kurum Adı:",
                "Firma Adı:",
                "Adı Soyadı:",
            ]

            for stop in stop_words:
                idx = value.lower().find(stop.lower())
                if idx > 0:
                    value = value[:idx].strip()

            return value

    return ""


def extract_contact(description, amount):
    if amount < 0:
        return extract_field(description, ["Alıcı", "Kurum Adı", "Firma Adı"])

    return extract_field(description, ["Gönderen", "Adı Soyadı"])


@app.get("/")
def health():
    return {"status": "ok"}


@app.post("/parse-bank-statement")
async def parse_bank_statement(request: Request):
    try:
        form = await request.form()

        uploaded_file = None

        for key in form:
            item = form[key]
            if isinstance(item, UploadFile) or isinstance(item, StarletteUploadFile):
                uploaded_file = item
                break

        if uploaded_file is None:
            return {
                "success": False,
                "error": "No uploaded file found",
                "lines": [],
            }

        content = uploaded_file.file.read()

        wb = load_workbook(BytesIO(content), data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))

        header_index = None
        headers = []

        for i, row in enumerate(rows):
            values = [str(c).strip() if c is not None else "" for c in row]
            joined = " ".join(values)

            if "İşlem Tarihi" in joined and "Açıklama" in joined and "Tutar" in joined:
                header_index = i
                headers = values
                break

        if header_index is None:
            return {
                "success": False,
                "error": "Header row not found",
                "lines": [],
            }

        def col(keyword):
            for idx, h in enumerate(headers):
                if keyword.lower() in h.lower():
                    return idx
            return -1

        date_col = col("İşlem Tarihi")
        desc_col = col("Açıklama")
        amount_col = col("Tutar")
        balance_col = col("Bakiye")
        ref_col = col("Referans")

        if date_col < 0 or desc_col < 0 or amount_col < 0:
            return {
                "success": False,
                "error": "Required columns not found",
                "headers": headers,
                "lines": [],
            }

        lines = []

        for row in rows[header_index + 1:]:
            if not row:
                continue

            if row[date_col] is None:
                continue

            desc = str(row[desc_col] or "").strip()
            amount = parse_amount(row[amount_col])
            balance = parse_amount(row[balance_col]) if balance_col >= 0 else 0
            ref = str(row[ref_col] or "").strip() if ref_col >= 0 else ""

            date_value = clean_date(row[date_col])
            contact_name = extract_contact(desc, amount)

            contact_type = "vendor" if amount < 0 else "customer"
            transaction_type = "vendor_payment" if amount < 0 else "customer_payment"

            unique_key = f"{ref}_{date_value}_{amount}"

            lines.append(
                {
                    "islem_tarihi": date_value,
                    "aciklama": desc,
                    "tutar": amount,
                    "bakiye": balance,
                    "referans_no": ref,
                    "contact_name": contact_name,
                    "contact_type": contact_type,
                    "transaction_type": transaction_type,
                    "unique_key": unique_key,
                }
            )

        return {
            "success": True,
            "count": len(lines),
            "lines": lines,
        }

    except Exception as e:
        return {
            "success": False,
            "error": str(e),
            "lines": [],
        }
